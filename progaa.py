from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.types import CallbackQuery
from config import TOKEN
import openpyxl
from openpyxl import load_workbook

workbook=load_workbook('bot.xlsx')
sheet=workbook['list1']

def metro(x,y):
    cell=sheet.cell(row=x,column=y).value
    return cell


bot = Bot(token=TOKEN)
dp = Dispatcher(bot)


@dp.message_handler(commands=['start','help'])
async def start_cmd_message(message: types.Message):
    await message.reply("Привет!\nНапиши /номер ветки метро, на которой\nты находишься.")
@dp.message_handler(commands=['1'])
async def vetka1(message: types.Message):
    await message.answer("Выбери станцию, на которой\nты находишься.\nЧтобы посмотреть фото парка, перейди по ссылке!", reply_markup=choice1)
@dp.message_handler(commands=['2'])
async def vetka2(message: types.Message):
    await message.answer("Выбери станцию, на которой\nты находишься.\nЧтобы посмотреть фото парка, перейди по ссылке!", reply_markup=choice2)
@dp.message_handler(commands=['3'])
async def vetka3(message: types.Message):
    await message.answer("Выбери станцию, на которой\nты находишься.\nЧтобы посмотреть фото парка, перейди по ссылке!", reply_markup=choice3)
@dp.message_handler(commands=['4'])
async def vetka4(message: types.Message):
    await message.answer("Выбери станцию, на которой\nты находишься.\nЧтобы посмотреть фото парка, перейди по ссылке!", reply_markup=choice4)
@dp.message_handler(commands=['5'])
async def vetka5(message: types.Message):
    await message.answer("Выбери станцию, на которой\nты находишься.\nЧтобы посмотреть фото парка, перейди по ссылке!", reply_markup=choice5)

#красная ветка метро - первая
choice1 = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            InlineKeyboardButton(text="Девяткино",callback_data="red1")
        ],
        [
            InlineKeyboardButton(text="Гражданский проспект", callback_data="red2")
        ],
        [
            InlineKeyboardButton(text="Академическая", callback_data="red3")
        ],
        [
            InlineKeyboardButton(text="Политехническая",callback_data="red4")
        ],

        [
            InlineKeyboardButton(text="Лесная", callback_data="red5")
        ],
        [
            InlineKeyboardButton(text="Выборгская", callback_data="red6")
        ],
        [
            InlineKeyboardButton(text="Площадь Ленина", callback_data="red7")
        ],
        [
            InlineKeyboardButton(text="Чернышевская", callback_data="red8")
        ],
        [
            InlineKeyboardButton(text="Площадь Восстания", callback_data="red9")
        ],
        [
            InlineKeyboardButton(text="Владимирская",  callback_data="red10")
        ],
        [
            InlineKeyboardButton(text="Пушкинская", callback_data="red11")
        ],
        [
            InlineKeyboardButton(text="Технологический институт", callback_data="red12")
        ],

        [
            InlineKeyboardButton(text="Балтийская", callback_data="red13")
        ],
        [
            InlineKeyboardButton(text="Нарвская", callback_data="red14")
        ],
        [
            InlineKeyboardButton(text="Кировский завод", callback_data="red15")
        ],
        [
            InlineKeyboardButton(text="Автово", callback_data="red16")
        ],

        [
            InlineKeyboardButton(text="Ленинский проспект", callback_data="red17")
        ],
        [
            InlineKeyboardButton(text="Проспект ветеранов", callback_data="red18")
        ]
    ]
)

@dp.callback_query_handler(text="red1")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(1,3))
    await callback.message.answer(metro(1,4))
    await callback.message.answer(metro(1,5))
@dp.callback_query_handler(text="red2")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(2, 3))
    await callback.message.answer(metro(2, 4))
    await callback.message.answer(metro(2, 5))
@dp.callback_query_handler(text="red3")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(3, 3))
    await callback.message.answer(metro(3, 4))
    await callback.message.answer(metro(3, 5))
    await callback.message.answer(metro(4, 3))
    await callback.message.answer(metro(4, 4))
    await callback.message.answer(metro(4, 5))
@dp.callback_query_handler(text="red4")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(5, 3))
    await callback.message.answer(metro(5, 4))
    await callback.message.answer(metro(5, 5))
@dp.callback_query_handler(text="red5")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(6, 3))
    await callback.message.answer(metro(6, 4))
    await callback.message.answer(metro(6, 5))
    await callback.message.answer(metro(7, 3))
    await callback.message.answer(metro(7, 4))
    await callback.message.answer(metro(7, 5))
    await callback.message.answer(metro(8, 3))
    await callback.message.answer(metro(8, 4))
    await callback.message.answer(metro(8, 5))
@dp.callback_query_handler(text="red6")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(9, 3))
    await callback.message.answer(metro(9, 4))
    await callback.message.answer(metro(9, 5))
@dp.callback_query_handler(text="red7")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(10, 3))
    await callback.message.answer(metro(10, 4))
    await callback.message.answer(metro(10, 5))
@dp.callback_query_handler(text="red8")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(11, 3))
    await callback.message.answer(metro(11, 4))
    await callback.message.answer(metro(11, 5))
    await callback.message.answer(metro(12, 3))
    await callback.message.answer(metro(12, 4))
    await callback.message.answer(metro(12, 5))
    await callback.message.answer(metro(13, 3))
    await callback.message.answer(metro(13, 4))
    await callback.message.answer(metro(13, 5))
    await callback.message.answer(metro(14, 3))
    await callback.message.answer(metro(14, 4))
    await callback.message.answer(metro(14, 5))
@dp.callback_query_handler(text="red9")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(15, 3))
    await callback.message.answer(metro(15, 4))
    await callback.message.answer(metro(15, 5))
    await callback.message.answer(metro(16, 3))
    await callback.message.answer(metro(16, 4))
    await callback.message.answer(metro(16, 5))
@dp.callback_query_handler(text="red10")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(17, 3))
    await callback.message.answer(metro(17, 4))
    await callback.message.answer(metro(17, 5))
    await callback.message.answer(metro(18, 3))
    await callback.message.answer(metro(18, 4))
    await callback.message.answer(metro(18, 5))
@dp.callback_query_handler(text="red11")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(19, 3))
    await callback.message.answer(metro(19, 4))
    await callback.message.answer(metro(19, 5))
@dp.callback_query_handler(text="red12")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(20, 3))
    await callback.message.answer(metro(20, 4))
    await callback.message.answer(metro(20, 5))
    await callback.message.answer(metro(21, 3))
    await callback.message.answer(metro(21, 4))
    await callback.message.answer(metro(21, 5))
@dp.callback_query_handler(text="red13")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(22, 3))
    await callback.message.answer(metro(22, 4))
    await callback.message.answer(metro(22, 5))
@dp.callback_query_handler(text="red14")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(23, 3))
    await callback.message.answer(metro(23, 4))
    await callback.message.answer(metro(23, 5))
@dp.callback_query_handler(text="red15")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(24, 3))
    await callback.message.answer(metro(24, 4))
    await callback.message.answer(metro(24, 5))
@dp.callback_query_handler(text="red16")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(25, 3))
    await callback.message.answer(metro(25, 4))
    await callback.message.answer(metro(25, 5))
@dp.callback_query_handler(text="red17")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(26, 3))
    await callback.message.answer(metro(26, 4))
    await callback.message.answer(metro(26, 5))
@dp.callback_query_handler(text="red18")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(27, 3))
    await callback.message.answer(metro(27, 4))
    await callback.message.answer(metro(27, 5))

#синия ветка метро - вторая
choice2 = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            InlineKeyboardButton(text="Парнас",callback_data="blue1")
        ],
        [
            InlineKeyboardButton(text="Проспект просвещения",callback_data="blue2")
        ],
        [
            InlineKeyboardButton(text="Озерки",callback_data="blue3")
        ],
        [
            InlineKeyboardButton(text="Удельная",callback_data="blue4")
        ],

        [
            InlineKeyboardButton(text="Пионерская",callback_data="blue5")
        ],
        [
            InlineKeyboardButton(text="Черная речка",callback_data="blue6")
        ],
        [
            InlineKeyboardButton(text="Петроградская",callback_data="blue7")
        ],
        [
            InlineKeyboardButton(text="Горьковская",callback_data="blue8")
        ],
        [
            InlineKeyboardButton(text="Невский проспект",callback_data="blue9")
        ],
        [
            InlineKeyboardButton(text="Сенная площадь", callback_data="blue10")
        ],
        [
            InlineKeyboardButton(text="Технологический институт", callback_data="blue11")
        ],
        [
            InlineKeyboardButton(text="Фрунзенская", callback_data="blue12")
        ],

        [
            InlineKeyboardButton(text="Московские ворота",callback_data="blue13")
        ],
        [
            InlineKeyboardButton(text="Электросила", callback_data="blue14")
        ],
        [
            InlineKeyboardButton(text="Парк Победы",callback_data="blue15")
        ],
        [
            InlineKeyboardButton(text="Московская",callback_data="blue16")
        ],

        [
            InlineKeyboardButton(text="Звездная",callback_data="blue17")
        ],
        [
            InlineKeyboardButton(text="Купчино", callback_data="blue18")
        ]
    ]
)
@dp.callback_query_handler(text="blue1")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(28,3))
    await callback.message.answer(metro(28,4))
    await callback.message.answer(metro(28,5))
@dp.callback_query_handler(text="blue2")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(29,3))
    await callback.message.answer(metro(29,4))
    await callback.message.answer(metro(29,5))
@dp.callback_query_handler(text="blue3")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(30,3))
    await callback.message.answer(metro(30,4))
    await callback.message.answer(metro(30,5))
@dp.callback_query_handler(text="blue4")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(31,3))
    await callback.message.answer(metro(31,4))
    await callback.message.answer(metro(31,5))
@dp.callback_query_handler(text="blue5")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(32,3))
    await callback.message.answer(metro(32,4))
    await callback.message.answer(metro(32,5))
@dp.callback_query_handler(text="blue6")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(33,3))
    await callback.message.answer(metro(33,4))
    await callback.message.answer(metro(33,5))
    await callback.message.answer(metro(34, 3))
    await callback.message.answer(metro(34, 4))
    await callback.message.answer(metro(34, 5))
@dp.callback_query_handler(text="blue7")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(35,3))
    await callback.message.answer(metro(35,4))
    await callback.message.answer(metro(35,5))
    await callback.message.answer(metro(36, 3))
    await callback.message.answer(metro(36, 4))
    await callback.message.answer(metro(36, 5))
    await callback.message.answer(metro(37, 3))
    await callback.message.answer(metro(37, 4))
    await callback.message.answer(metro(37, 5))
@dp.callback_query_handler(text="blue8")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(38,3))
    await callback.message.answer(metro(38,4))
    await callback.message.answer(metro(38,5))
    await callback.message.answer(metro(39, 3))
    await callback.message.answer(metro(39, 4))
    await callback.message.answer(metro(39, 5))
    await callback.message.answer(metro(40, 3))
    await callback.message.answer(metro(40, 4))
    await callback.message.answer(metro(40, 5))
@dp.callback_query_handler(text="blue9")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(41,3))
    await callback.message.answer(metro(41,4))
    await callback.message.answer(metro(41,5))
    await callback.message.answer(metro(42, 3))
    await callback.message.answer(metro(42, 4))
    await callback.message.answer(metro(42, 5))
    await callback.message.answer(metro(43, 3))
    await callback.message.answer(metro(43, 4))
    await callback.message.answer(metro(43, 5))
    await callback.message.answer(metro(44, 3))
    await callback.message.answer(metro(44, 4))
    await callback.message.answer(metro(44, 5))
@dp.callback_query_handler(text="blue10")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(45, 3))
    await callback.message.answer(metro(45, 4))
    await callback.message.answer(metro(45, 5))
    await callback.message.answer(metro(46, 3))
    await callback.message.answer(metro(46, 4))
    await callback.message.answer(metro(46, 5))
@dp.callback_query_handler(text="blue11")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(47, 3))
    await callback.message.answer(metro(47, 4))
    await callback.message.answer(metro(47, 5))
    await callback.message.answer(metro(48, 3))
    await callback.message.answer(metro(48, 4))
    await callback.message.answer(metro(48, 5))
@dp.callback_query_handler(text="blue12")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(49, 3))
    await callback.message.answer(metro(49, 4))
    await callback.message.answer(metro(49, 5))
    await callback.message.answer(metro(50, 3))
    await callback.message.answer(metro(50, 4))
    await callback.message.answer(metro(50, 5))
@dp.callback_query_handler(text="blue13")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(51, 3))
    await callback.message.answer(metro(51, 4))
    await callback.message.answer(metro(51, 5))
@dp.callback_query_handler(text="blue14")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(52, 3))
    await callback.message.answer(metro(52, 4))
    await callback.message.answer(metro(52, 5))
@dp.callback_query_handler(text="blue15")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(53, 3))
    await callback.message.answer(metro(53, 4))
    await callback.message.answer(metro(53, 5))
    await callback.message.answer(metro(54, 3))
    await callback.message.answer(metro(54, 4))
    await callback.message.answer(metro(54, 5))
    await callback.message.answer(metro(55, 3))
    await callback.message.answer(metro(55, 4))
    await callback.message.answer(metro(55, 5))
@dp.callback_query_handler(text="blue16")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(56, 3))
    await callback.message.answer(metro(56, 4))
    await callback.message.answer(metro(56, 5))
@dp.callback_query_handler(text="blue17")
async def red2(callback: types.CallbackQuery):
    await callback.message.answer(metro(56, 3))
    await callback.message.answer(metro(56, 4))
    await callback.message.answer(metro(56, 5))
@dp.callback_query_handler(text="blue18")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(57, 3))
    await callback.message.answer(metro(57, 4))
    await callback.message.answer(metro(57, 5))

#зеленая ветка метро - третья
choice3 = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            InlineKeyboardButton(text="Беговая",callback_data="green1")
        ],
        [
            InlineKeyboardButton(text="Зенит",callback_data="green2")
        ],
        [
            InlineKeyboardButton(text="Приморская",callback_data="green3")
        ],
        [
            InlineKeyboardButton(text="Василеостровская",callback_data="green4")
        ],

        [
            InlineKeyboardButton(text="Гостиный двор",callback_data="green5")
        ],
        [
            InlineKeyboardButton(text="Маяковская",callback_data="green6")
        ],
        [
            InlineKeyboardButton(text="Площадь Александра Невского",callback_data="green7")
        ],
        [
            InlineKeyboardButton(text="Елизаровская",callback_data="green8")
        ],
        [
            InlineKeyboardButton(text="Ломоносовская",callback_data="green9")
        ],
        [
            InlineKeyboardButton(text="Пролетарская",callback_data="green10")
        ],
        [
            InlineKeyboardButton(text="Обухово", callback_data="green11")
        ],
        [
            InlineKeyboardButton(text="Рыбацкое",callback_data="green12")
        ]
    ]
)
@dp.callback_query_handler(text="green1")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(59,3))
    await callback.message.answer(metro(59,4))
    await callback.message.answer(metro(59,5))
@dp.callback_query_handler(text="green2")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(60,3))
    await callback.message.answer(metro(60,4))
    await callback.message.answer(metro(60,5))
@dp.callback_query_handler(text="green3")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(61, 3))
    await callback.message.answer(metro(61, 4))
    await callback.message.answer(metro(61, 5))
    await callback.message.answer(metro(62, 3))
    await callback.message.answer(metro(62, 4))
    await callback.message.answer(metro(62, 5))
    await callback.message.answer(metro(63, 3))
    await callback.message.answer(metro(63, 4))
    await callback.message.answer(metro(63, 5))
@dp.callback_query_handler(text="green4")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(64, 3))
    await callback.message.answer(metro(64, 4))
    await callback.message.answer(metro(64, 5))
    await callback.message.answer(metro(65, 3))
    await callback.message.answer(metro(65, 4))
    await callback.message.answer(metro(65, 5))
    await callback.message.answer(metro(66, 3))
    await callback.message.answer(metro(66, 4))
    await callback.message.answer(metro(66, 5))
@dp.callback_query_handler(text="green5")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(67, 3))
    await callback.message.answer(metro(67, 4))
    await callback.message.answer(metro(67, 5))
    await callback.message.answer(metro(68, 3))
    await callback.message.answer(metro(68, 4))
    await callback.message.answer(metro(68, 5))
    await callback.message.answer(metro(69, 3))
    await callback.message.answer(metro(69, 4))
    await callback.message.answer(metro(69, 5))
    await callback.message.answer(metro(70, 3))
    await callback.message.answer(metro(70, 4))
    await callback.message.answer(metro(70, 5))
@dp.callback_query_handler(text="green6")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(71, 3))
    await callback.message.answer(metro(71, 4))
    await callback.message.answer(metro(71, 5))
    await callback.message.answer(metro(72, 3))
    await callback.message.answer(metro(72, 4))
    await callback.message.answer(metro(72, 5))
@dp.callback_query_handler(text="green7")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(73, 3))
    await callback.message.answer(metro(73, 4))
    await callback.message.answer(metro(73, 5))
    await callback.message.answer(metro(74, 3))
    await callback.message.answer(metro(74, 4))
    await callback.message.answer(metro(74, 5))
@dp.callback_query_handler(text="green8")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(75, 3))
    await callback.message.answer(metro(75, 4))
    await callback.message.answer(metro(75, 5))
    await callback.message.answer(metro(76, 3))
    await callback.message.answer(metro(76, 4))
    await callback.message.answer(metro(76, 5))
@dp.callback_query_handler(text="green9")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(77, 3))
    await callback.message.answer(metro(77, 4))
    await callback.message.answer(metro(77, 5))
    await callback.message.answer(metro(78, 3))
    await callback.message.answer(metro(78, 4))
    await callback.message.answer(metro(78, 5))
@dp.callback_query_handler(text="green10")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(79, 3))
    await callback.message.answer(metro(79, 4))
    await callback.message.answer(metro(79, 5))
    await callback.message.answer(metro(80, 3))
    await callback.message.answer(metro(80, 4))
    await callback.message.answer(metro(80, 5))
@dp.callback_query_handler(text="green11")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(81, 3))
    await callback.message.answer(metro(81, 4))
    await callback.message.answer(metro(81, 5))
@dp.callback_query_handler(text="green12")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(82, 3))
    await callback.message.answer(metro(82, 4))
    await callback.message.answer(metro(82, 5))
#оранжевая ветка метро - четвертая
choice4 = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            InlineKeyboardButton(text="Спасская",callback_data="orange1")
        ],
        [
            InlineKeyboardButton(text="Достоевская",callback_data="orange2")
        ],
        [
            InlineKeyboardButton(text="Лиговский проспект",callback_data="orange3")
        ],
        [
            InlineKeyboardButton(text="Площадь Александра Невского",callback_data="orange4")
        ],

        [
            InlineKeyboardButton(text="Новочеркасская",callback_data="orange5")
        ],
        [
            InlineKeyboardButton(text="Ладожская",callback_data="orange6")
        ],
        [
            InlineKeyboardButton(text="Проспект Большевиков",callback_data="orange7")
        ],
        [
            InlineKeyboardButton(text="Улица Дыбенко",callback_data="orange8")
        ]
    ]
)
@dp.callback_query_handler(text="orange1")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(83, 3))
    await callback.message.answer(metro(83, 4))
    await callback.message.answer(metro(83, 5))
    await callback.message.answer(metro(84, 3))
    await callback.message.answer(metro(84, 4))
    await callback.message.answer(metro(84, 5))
@dp.callback_query_handler(text="orange2")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(85, 3))
    await callback.message.answer(metro(85, 4))
    await callback.message.answer(metro(85, 5))
    await callback.message.answer(metro(86, 3))
    await callback.message.answer(metro(86, 4))
    await callback.message.answer(metro(86, 5))
@dp.callback_query_handler(text="orange3")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(87, 3))
    await callback.message.answer(metro(87, 4))
    await callback.message.answer(metro(87, 5))
@dp.callback_query_handler(text="orange4")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(88, 3))
    await callback.message.answer(metro(88, 4))
    await callback.message.answer(metro(88, 5))
    await callback.message.answer(metro(89, 3))
    await callback.message.answer(metro(89, 4))
    await callback.message.answer(metro(89, 5))
@dp.callback_query_handler(text="orange5")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(90, 3))
    await callback.message.answer(metro(90, 4))
    await callback.message.answer(metro(90, 5))
    await callback.message.answer(metro(91, 3))
    await callback.message.answer(metro(91, 4))
    await callback.message.answer(metro(91, 5))
    await callback.message.answer(metro(92, 3))
    await callback.message.answer(metro(92, 4))
    await callback.message.answer(metro(92, 5))
@dp.callback_query_handler(text="orange6")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(93, 3))
    await callback.message.answer(metro(93, 4))
    await callback.message.answer(metro(93, 5))
    await callback.message.answer(metro(94, 3))
    await callback.message.answer(metro(94, 4))
    await callback.message.answer(metro(94, 5))
@dp.callback_query_handler(text="orange7")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(95, 3))
    await callback.message.answer(metro(95, 4))
    await callback.message.answer(metro(95, 5))
@dp.callback_query_handler(text="orange8")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(96, 3))
    await callback.message.answer(metro(96, 4))
    await callback.message.answer(metro(96, 5))
    await callback.message.answer(metro(97, 3))
    await callback.message.answer(metro(97, 4))
    await callback.message.answer(metro(97, 5))
    await callback.message.answer(metro(98, 3))
    await callback.message.answer(metro(98, 4))
    await callback.message.answer(metro(98, 5))
    await callback.message.answer(metro(99, 3))
    await callback.message.answer(metro(99, 4))
    await callback.message.answer(metro(99, 5))
#фиолетовая ветка метро - пятая
choice5 = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            InlineKeyboardButton(text="Комендатский проспект",callback_data="pink1")
        ],
        [
            InlineKeyboardButton(text="Старая деревня",callback_data="pink2")
        ],
        [
            InlineKeyboardButton(text="Крестовский остров",callback_data="pink3")
        ],
        [
            InlineKeyboardButton(text="Чкаловская",callback_data="pink4")
        ],

        [
            InlineKeyboardButton(text="Спортивная",callback_data="pink5")
        ],
        [
            InlineKeyboardButton(text="Адмиралтейская",callback_data="pink6")
        ],
        [
            InlineKeyboardButton(text="Садовая",callback_data="pink7")
        ],
        [
            InlineKeyboardButton(text="Звенигородская ",callback_data="pink8")
        ],
        [
            InlineKeyboardButton(text="Обводный канал",callback_data="pink9")
        ],
        [
            InlineKeyboardButton(text="Волковская", callback_data="pink10")
        ],
        [
            InlineKeyboardButton(text="Бухаресткая", callback_data="pink11")
        ],
        [
            InlineKeyboardButton(text="Международная",callback_data="pink12")
        ],

        [
            InlineKeyboardButton(text="Проспект славы", callback_data="pink13")
        ],
        [
            InlineKeyboardButton(text="Дунайская",callback_data="pink14")
        ]
    ]
)
@dp.callback_query_handler(text="pink1")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(100, 3))
    await callback.message.answer(metro(100, 4))
    await callback.message.answer(metro(100, 5))
    await callback.message.answer(metro(101, 3))
    await callback.message.answer(metro(101, 4))
    await callback.message.answer(metro(101, 5))
@dp.callback_query_handler(text="pink2")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(102, 3))
    await callback.message.answer(metro(102, 4))
    await callback.message.answer(metro(102, 5))
@dp.callback_query_handler(text="pink3")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(103, 3))
    await callback.message.answer(metro(103, 4))
    await callback.message.answer(metro(103, 5))
    await callback.message.answer(metro(104, 3))
    await callback.message.answer(metro(104, 4))
    await callback.message.answer(metro(104, 5))
@dp.callback_query_handler(text="pink4")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(105, 3))
    await callback.message.answer(metro(105, 4))
    await callback.message.answer(metro(105, 5))
    await callback.message.answer(metro(106, 3))
    await callback.message.answer(metro(106, 4))
    await callback.message.answer(metro(106, 5))
    await callback.message.answer(metro(107, 3))
    await callback.message.answer(metro(107, 4))
    await callback.message.answer(metro(107, 5))
    await callback.message.answer(metro(108, 3))
    await callback.message.answer(metro(108, 4))
    await callback.message.answer(metro(108, 5))
@dp.callback_query_handler(text="pink5")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(109, 3))
    await callback.message.answer(metro(109, 4))
    await callback.message.answer(metro(109, 5))
@dp.callback_query_handler(text="pink6")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(110, 3))
    await callback.message.answer(metro(110, 4))
    await callback.message.answer(metro(110, 5))
    await callback.message.answer(metro(111, 3))
    await callback.message.answer(metro(111, 4))
    await callback.message.answer(metro(111, 5))
    await callback.message.answer(metro(112, 3))
    await callback.message.answer(metro(112, 4))
    await callback.message.answer(metro(112, 5))
    await callback.message.answer(metro(113, 3))
    await callback.message.answer(metro(113, 4))
    await callback.message.answer(metro(113, 5))
@dp.callback_query_handler(text="pink7")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(114, 3))
    await callback.message.answer(metro(114, 4))
    await callback.message.answer(metro(114, 5))
    await callback.message.answer(metro(115, 3))
    await callback.message.answer(metro(115, 4))
    await callback.message.answer(metro(115, 5))
@dp.callback_query_handler(text="pink8")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(116, 3))
    await callback.message.answer(metro(116, 4))
    await callback.message.answer(metro(116, 5))
@dp.callback_query_handler(text="pink9")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(117, 3))
    await callback.message.answer(metro(117, 4))
    await callback.message.answer(metro(117, 5))
@dp.callback_query_handler(text="pink10")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(118, 3))
    await callback.message.answer(metro(118, 4))
    await callback.message.answer(metro(118, 5))
@dp.callback_query_handler(text="pink11")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(119, 3))
    await callback.message.answer(metro(119, 4))
    await callback.message.answer(metro(119, 5))
    await callback.message.answer(metro(120, 3))
    await callback.message.answer(metro(120, 4))
    await callback.message.answer(metro(120, 5))
@dp.callback_query_handler(text="pink12")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(121, 3))
    await callback.message.answer(metro(121, 4))
    await callback.message.answer(metro(121, 5))
    await callback.message.answer(metro(122, 3))
    await callback.message.answer(metro(122, 4))
    await callback.message.answer(metro(122, 5))
@dp.callback_query_handler(text="pink13")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(123, 3))
    await callback.message.answer(metro(123, 4))
    await callback.message.answer(metro(123, 5))
@dp.callback_query_handler(text="pink14")
async def red2(callback : types.CallbackQuery):
    await callback.message.answer(metro(124, 3))
    await callback.message.answer(metro(124, 4))
    await callback.message.answer(metro(124, 5))
    await callback.message.answer(metro(125, 3))
    await callback.message.answer(metro(125, 4))
    await callback.message.answer(metro(125, 5))













if __name__ == '__main__': #был ли файл запущен напрямую
    executor.start_polling(dp)








